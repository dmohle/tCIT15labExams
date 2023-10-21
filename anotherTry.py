import openpyxl
from datetime import date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
# from openpyxl.drawing.chart import PieChart, Reference
from openpyxl.styles import NamedStyle
import datetime

# Define the directory and file paths
directory = r"C:\2023spring\pythonRoot\lastExcelTry"
input_file = directory + r"\ExcelLabTestStart.xlsx"
output_file = directory + r"\CIT15 Excel lab Test_Final_End.xlsx"

# Load the existing workbook
wb = openpyxl.load_workbook(input_file)

# Access the 'Session1' worksheet
if 'Session1' in wb.sheetnames:
    ws = wb['Session1']
else:
    ws = wb.create_sheet('Session1')

# Enter "Start Date" in cell A17 and insert the current date into cell B17
ws['A17'] = 'Start Date'
ws['B17'] = date.today()

# Format cells C7:C11 to display as percentages
for cell in ws['C7:C11']:
    cell[0].number_format = '0.00%'

# Step 4
# On the Session1 worksheet, in cell E7, create a formula
# that will calculate the Session Fee with Member Discount using D7 times C7.
ws['E7'] = '=D7*C7'

# Step 5
# On the Session1 worksheet, change the value in cell D8 to 280,
# and then delete the contents in cell A13.
ws['D8'] = 280
ws['A13'].value = None

# Step 6
# On the Session1 worksheet, in cell H7, create a formula that
# will calculate the Total Fees collected for the Introduction to Computer Literacy class.
ws['H7'] = '=F7*D7 + G7*E7'

# Step 7
for row in range(8, 12):  # From E8:E11 and H8:H11
    ws[f'E{row}'] = f'={get_column_letter(4)}{row}*{get_column_letter(3)}{row}'  # Copy formula from E7
    ws[f'H{row}'] = f'=(F{row}*D{row})+(G{row}*E{row})'  # Copy formula from H7

# Step 8
if 'Sheet1' in wb.sheetnames:
    wb.remove(wb['Sheet1'])

# Step 9
ws.insert_rows(5)
ws.delete_rows(3)

# Step 10
copy_range = ws['A5:E11']
if 'Class Fees' not in wb.sheetnames:
    class_fees_ws = wb.create_sheet('Class Fees')
else:
    class_fees_ws = wb['Class Fees']

for row_data, row_cells in enumerate(copy_range, start=4):
    for col_data, cell in enumerate(row_cells, start=1):
        class_fees_ws[get_column_letter(col_data) + str(row_data)].value = cell.value

# Step 11
class_fees_ws.column_dimensions['A'].width = 30
class_fees_ws.column_dimensions['D'].width = 14
class_fees_ws.column_dimensions['E'].width = 14

# Step 12
copy_range_values = ws['A1:H2']
for row_data, row_cells in enumerate(copy_range_values, start=1):
    for col_data, cell in enumerate(row_cells, start=1):
        class_fees_ws[get_column_letter(col_data) + str(row_data)].value = cell.value

# Step 13
class_fees_ws.merge_cells('A1:E1')
class_fees_ws.merge_cells('A2:E2')
class_fees_ws['A1'].alignment = Alignment(horizontal='center')
class_fees_ws['A2'].alignment = Alignment(horizontal='center')

# Step 14
thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))
tan_fill = PatternFill(start_color="FFD2B48C", end_color="FFD2B48C", fill_type="solid")

for row in class_fees_ws['A4:E5']:
    for cell in row:
        cell.border = thick_border
        cell.fill = tan_fill

# Step 15
for row in class_fees_ws['D6:E10']:
    for cell in row:
        cell.number_format = "$#,##0"

# Step 16
class_fees_ws.page_setup.centerHorizontally = True
# check this to be sure this works!
class_fees_ws.HeaderFooter.rightHeader = "Fresno City College"

# Step 17: Update Font Sizes and Styles for Session1 A1 and A2
ws['A1'].font = Font(size=14, bold=True)
ws['A2'].font = Font(size=12, bold=True)
ws.merge_cells('A1:H1')
ws.merge_cells('A2:H2')

# Step 18: Apply accounting style to specific ranges in Session1
for row in range(7, 12):
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'E{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].number_format = '#,##0.00'

# Step 19: Bold and Fill Color for range A5:H6
bold_font = Font(bold=True)
fill_color = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid")  # Olive Green, Accent 3
for col in range(1, 9):
    for row in [5, 6]:
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.font = bold_font
        cell.fill = fill_color

# Step 20: Input and formulas in Session1 C13, D13 and E13:H13
ws['C13'] = 'Total'
ws['C14'] = 'Average'
ws['D13'] = f'=SUM(D7:D11)'
ws['E13'] = f'=SUM(E7:E11)'
# Copy formulas for F13:H13
for col in ['F', 'G', 'H']:
    ws[f'{col}13'].value = f'=SUM({col}7:{col}11)'

# Step 21: AVERAGE formulas in Session1 D14:H14 and format adjustments
ws['D14'] = f'=AVERAGE(D7:D11)'
# Copy formulas for E14:H14
for col in ['E', 'F', 'G', 'H']:
    ws[f'{col}14'].value = f'=AVERAGE({col}7:{col}11)'

# Adjust number format for F13:G14
for col in ['F', 'G']:
    for row in [13, 14]:
        ws[f'{col}{row}'].number_format = '#,##0'

# Step 22: Session1 worksheet - Landscape orientation, header with Current Date & "Fresno City College"
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.HeaderFooter.right_header = datetime.datetime.now().strftime('%Y-%m-%d')
ws.HeaderFooter.right_header = "Fresno City College"

# Step 23: Insert 3-D Pie chart in Session1 worksheet
# can't do charts -- do this manually

# chart = PieChart()
#  chart.add_data(data, titles_from_data=True)
#  chart.set_categories(labels)
#  chart.title = "Class Enrollment"
#  chart.style = 3
#  ws.add_chart(chart, "A19")

# Step 24: Format the pie chart to Style 3 and set title
# This step is essentially handled in the previous step by setting chart.style and chart.title

# Step 25: Change Orientation of Class Fees worksheet to Landscape
class_fees_ws.page_setup.orientation = class_fees_ws.ORIENTATION_LANDSCAPE

# Step 26: Save and close the document. Exit Excel.



# Save the workbook with a new name
wb.save(output_file)

print(f"File saved as {output_file}")
